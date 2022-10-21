from distutils.log import error
import glob
import os
import string
import sys
from warnings import catch_warnings

import pandas as pd
from openpyxl import load_workbook

"""
Author: Motlatsi
Created: 14/Oct/22

The script's main intention/function is to update/format the subnet works from the uGridNet output and the corresponding
branches.
    
The Other function of the script is to return the Total number of poles and their types classified as LV and MV
"""

class Modify_uGridNet_excel:
    
    def __init__(self, NetworkLength_df, DropLines_df, excel_file):
        self.NetworkLength_df = NetworkLength_df
        self.DropLines_df = DropLines_df
        self.excel_file = excel_file   
        Modify_uGridNet_excel.Modify_Subnetwork(self)
        
    def Count_Poles(self):
        MV_Pole = 0
        PoleClass_df = pd.read_excel(self.excel_file,index_col=0, sheet_name="PoleClasses")
        self.Pole_Length = len(PoleClass_df)
        for Pole in PoleClass_df["Type"]:
            if "MV" in Pole:
                MV_Pole +=1
        count_pole_info = f"Pole infomation: LV pole:{self.Pole_Length-MV_Pole}  MV pole:{MV_Pole}"
        return (count_pole_info)
    
    def File_Directies():
        
        path=os.getcwd()
        files_in_cdir = os.listdir(path)
        consession_files_path = []
        Village_name_list = []
        consession_excels = []
    
        for vill_name in files_in_cdir:  
            if "Archive" not in vill_name and "SC" not in vill_name and ".py" not in vill_name: 
                SOURCE_DIR = vill_name
                files_in_odir = glob.glob(SOURCE_DIR )
                if ('Arc') not in files_in_odir and ('Ach') not in files_in_odir and ('GIS') not in files_in_odir and ("xlsx") not in files_in_odir and ("py") not in files_in_odir:
                    Village_name_list.append(files_in_cdir)
                    for inner_file in files_in_odir: #inner_file is the village_name folder
                        for Infile in glob.glob(inner_file + "/" + inner_file ): #concatinatinating inner_file twice takes us to the destination dir
                            destination_dir = os.listdir(Infile)
                            for file in destination_dir:
                               if ('Arc') not in file and ('Ach') not in file and ('GIS') not in file and ('uGridNet') in file and file.endswith('xlsx'):
                                    consession_excels.append(file)
                                    consession_files_path.append(f"{os.path.abspath(Infile)}\\{file}")
        
        Modify_uGridNet_excel.instantiate(consession_files_path, consession_excels)
        
    def instantiate(consession_files_path,consession_excels):
        
        for file in (consession_files_path):
            excel_file = file
            try:
                if (os.path.exists(excel_file)):
                    network_length_df = pd.read_excel(excel_file, index_col = 0, sheet_name="NetworkLength")
                    DropLines_df = pd.read_excel(excel_file, index_col = 0, sheet_name="DropLines") 
                    #print(network_length_df)
                    Modify_uGridNet_excel(network_length_df, DropLines_df,excel_file)                             
                else:
                    print(f"No such file path as {excel_file}")
            except:
                print ("Failed: ", error)                               
           
    def Update_NetworkLength(self,New_Branch,New_subnetwork, idx, Type):
        self.wb = load_workbook(self.excel_file)
        self.ws = self.wb["NetworkLength"]
        self.wb.save(self.excel_file)
        
        MV_pole_count = 0
    
        if "MV" in Type:
            MV_pole_count+=1
            self.ws.cell(row=idx+2, column = 9, value = New_subnetwork)
            self.wb.save(self.excel_file)      
        else:  
            self.ws.cell(row=idx+2, column = 9, value = New_subnetwork)
            self.ws.cell(row=idx+2, column = 10, value = New_Branch)
            self.wb.save(self.excel_file)
   
    def Update_DropLines(self,New_subnetwork, idx):

        self.wb = load_workbook(self.excel_file)
        self.ws=self.wb["DropLines"]
        self.ws.cell(row=idx+2, column = 7, value = New_subnetwork)
        self.wb.save(self.excel_file)
             
    def Modify_Subnetwork(self):
        
        print(f"\n\nNow on {self.excel_file}\nUpdating NetworkLength.....")
        Branch_Letters = (string.ascii_uppercase)
        for idx in range(len(self.NetworkLength_df)):
            Type = self.NetworkLength_df.iat[idx,4]
            Pole_ID_From = self.NetworkLength_df.iat[idx,5]
            Pole_ID_To = self.NetworkLength_df.iat[idx,6]
            SubNetwork = self.NetworkLength_df.iat[idx,7]
            
            count = 0
            for i in range(len(Pole_ID_From)):
                if "_" in Pole_ID_From[i] and "LV" in Type:
                    count += 1
                    if count > 1:    
                        if "M" in Pole_ID_From and Pole_ID_From[-1] not in Branch_Letters:
                            New_subnetwork = f"{SubNetwork[0:2]}{Pole_ID_To[i+1]}" 
                            New_Branch = f"{Pole_ID_To[i+2]}" 
                        if "M" in Pole_ID_From and Pole_ID_From[-1] in Branch_Letters:    
                             
                            New_Branch = f"{Pole_ID_To[i+2]}"                              
                            Modify_uGridNet_excel.Update_NetworkLength(self,New_Branch,New_subnetwork, idx, Type)                                
                        else:
                            New_subnetwork = f"{SubNetwork[0:2]}{Pole_ID_From[i+1]}" 
                            New_Branch = f"{Pole_ID_From[i+2]}"
                            
                            Modify_uGridNet_excel.Update_NetworkLength(self,New_Branch,New_subnetwork, idx, Type) 
                            
                        count = 0 
                         
            if "MV" in Type:
                New_subnetwork = f"{Pole_ID_From[4:6]}M"
                New_Branch=""
                Modify_uGridNet_excel.Update_NetworkLength(self,New_Branch, New_subnetwork, idx,Type)
                
        Modify_uGridNet_excel.Count_Poles(self) 
              
        print("Updating Droplines...")
        for idx in range(len(self.DropLines_df)):
            DropPoleID = self.DropLines_df.iat[idx,4]
            New_subnetwork = f"{SubNetwork[0:2]}{DropPoleID[7]}"
            
            Modify_uGridNet_excel.Update_DropLines(self, New_subnetwork, idx)
            if len(self.DropLines_df) == idx+1:
                self.wb.save(self.excel_file)
                Modify_uGridNet_excel.Update_NetworkLength
                
        print(Modify_uGridNet_excel.Count_Poles(self), f" Total Poles {self.Pole_Length}\nDone!!!")
                

if __name__ == "__main__":
    
    Modify_uGridNet_excel.File_Directies()
    

